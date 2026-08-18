"""Geração da aba RANKING de saldos críticos."""

from collections.abc import Iterable

from openpyxl import Workbook

from app.domain import RegistroFuncionario, formatar_horas
from app.reports.styles import criar_estilos_resumo


def criar_aba_ranking(
    wb: Workbook,
    dados: Iterable[RegistroFuncionario],
) -> None:
    """Cria o ranking de saldos abaixo de -8h e acima de +8h."""
    if "RANKING" in wb.sheetnames:
        del wb["RANKING"]

    ws = wb.create_sheet("RANKING")
    estilos = criar_estilos_resumo()

    limite_8h_min = 8 * 60
    registros = list(dados)

    # Não altera a forma de cálculo: usa o saldo já calculado em minutos.
    # A aba lista todos os casos críticos, mantendo a ordem crescente atual.
    negativos = sorted(
        [
            registro
            for registro in registros
            if registro.banco_saldo_minutos < -limite_8h_min
        ],
        key=lambda registro: registro.banco_saldo_minutos,
    )
    positivos = sorted(
        [
            registro
            for registro in registros
            if registro.banco_saldo_minutos > limite_8h_min
        ],
        key=lambda registro: registro.banco_saldo_minutos,
    )

    ws["A1"] = "DEVEDORES - ABAIXO DE -8 HORAS"
    ws["A1"].font = estilos["titulo"]
    ws.append(["Funcionário", "Número de matrícula", "Departamento", "Banco Saldo"])
    for registro in negativos:
        ws.append(
            [
                registro.nome,
                registro.matricula,
                registro.departamento,
                formatar_horas(registro.banco_saldo_minutos),
            ]
        )

    inicio_segunda_secao = ws.max_row + 3
    ws.cell(row=inicio_segunda_secao, column=1, value="HORAS EXTRAS - ACIMA DE 8 HORAS")
    ws.cell(row=inicio_segunda_secao, column=1).font = estilos["titulo"]
    ws.append(["Funcionário", "Número de matrícula", "Departamento", "Banco Saldo"])
    for registro in positivos:
        ws.append(
            [
                registro.nome,
                registro.matricula,
                registro.departamento,
                formatar_horas(registro.banco_saldo_minutos),
            ]
        )

    for row in (2, inicio_segunda_secao + 1):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = estilos["cabecalho_font"]
            cell.fill = estilos["cabecalho_fill"]
            cell.alignment = estilos["centro"]
            cell.border = estilos["borda"]

    for row in range(3, ws.max_row + 1):
        if row == inicio_segunda_secao:
            continue
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = estilos["borda"]
            if col in (1, 3):
                cell.alignment = estilos["esquerda"]
            elif col == 2:
                cell.alignment = estilos["centro"]
                cell.number_format = "@"
            else:
                cell.alignment = estilos["direita"]

    ws.freeze_panes = None
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 16
