"""Geração da aba RESUMO por departamento."""

from collections import OrderedDict
from collections.abc import Iterable

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.domain import RegistroFuncionario, formatar_horas
from app.reports.styles import criar_estilos_resumo


def criar_aba_resumo(
    wb: Workbook,
    dados: Iterable[RegistroFuncionario],
) -> None:
    """Cria o resumo dos saldos já calculados, agrupado por departamento."""
    if "RESUMO" in wb.sheetnames:
        del wb["RESUMO"]

    ws = wb.create_sheet("RESUMO")
    estilos = criar_estilos_resumo()

    limite_8h_min = 8 * 60
    vermelho = PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")
    amarelo = PatternFill(fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00")

    resumo = OrderedDict()
    for registro in sorted(
        dados,
        key=lambda item: str(item.departamento or "SEM DEPARTAMENTO").lower(),
    ):
        departamento = (
            registro.departamento
            if registro.departamento not in (None, "")
            else "SEM DEPARTAMENTO"
        )
        resumo.setdefault(departamento, 0)
        resumo[departamento] += registro.banco_saldo_minutos

    # Não altera o cálculo: apenas ordena o saldo já totalizado em ordem crescente.
    itens_resumo = sorted(resumo.items(), key=lambda item: item[1])

    ws["A1"] = "Resumo por departamento"
    ws["A1"].font = estilos["titulo"]
    ws.merge_cells("A1:C1")
    ws["A1"].alignment = estilos["esquerda"]

    ws.append(["Departamento", "Horas", "Horas_num"])

    for departamento, total_min in itens_resumo:
        ws.append([departamento, formatar_horas(total_min), total_min / 60])

    linha_total = ws.max_row + 1
    total_geral_min = sum(resumo.values())
    ws.cell(row=linha_total, column=1, value="TOTAL")
    ws.cell(row=linha_total, column=2, value=formatar_horas(total_geral_min))
    ws.cell(row=linha_total, column=3, value=total_geral_min / 60)

    for col in range(1, 4):
        cell = ws.cell(row=2, column=col)
        cell.font = estilos["cabecalho_font"]
        cell.fill = estilos["cabecalho_fill"]
        cell.alignment = estilos["centro"]
        cell.border = estilos["borda"]

    for row in range(3, linha_total + 1):
        total_min = ws.cell(row=row, column=3).value
        fill = None
        if isinstance(total_min, (int, float)):
            total_min = total_min * 60
            if total_min < -limite_8h_min:
                fill = vermelho
            elif total_min > limite_8h_min:
                fill = amarelo

        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.border = estilos["borda"]
            cell.alignment = estilos["esquerda"] if col == 1 else estilos["direita"]
            if fill and row != linha_total:
                cell.fill = fill

    for col in range(1, 4):
        cell = ws.cell(row=linha_total, column=col)
        cell.font = estilos["linha_total_font"]
        cell.fill = estilos["linha_total_fill"]
        cell.border = estilos["borda"]

    linha_legenda = linha_total + 3
    ws.cell(row=linha_legenda, column=1, value="LEGENDA DE CORES")
    ws.cell(row=linha_legenda, column=1).font = estilos["titulo"]

    ws.cell(row=linha_legenda + 1, column=1, value="Devedores")
    ws.cell(row=linha_legenda + 1, column=2, value="Saldo menor que -8h")
    ws.cell(row=linha_legenda + 1, column=1).fill = vermelho

    ws.cell(row=linha_legenda + 2, column=1, value="Extras")
    ws.cell(row=linha_legenda + 2, column=2, value="Saldo maior que 8h")
    ws.cell(row=linha_legenda + 2, column=1).fill = amarelo

    for row in range(linha_legenda, linha_legenda + 3):
        for col in range(1, 3):
            cell = ws.cell(row=row, column=col)
            cell.border = estilos["borda"]
            cell.alignment = (
                estilos["centro"] if row == linha_legenda else estilos["esquerda"]
            )

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:B{linha_total}"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].hidden = True
