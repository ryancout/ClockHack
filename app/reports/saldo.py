"""Geração da aba consolidada SALDO."""

from collections.abc import Iterable

from openpyxl import Workbook

from app.domain import RegistroFuncionario, formatar_horas
from app.reports.styles import criar_estilos_resumo


def criar_aba_saldo(
    wb: Workbook,
    dados: Iterable[RegistroFuncionario],
) -> None:
    """Cria uma visão consolidada de saldo sem recalcular valores de horas."""
    if "SALDO" in wb.sheetnames:
        del wb["SALDO"]

    ws = wb.create_sheet("SALDO")
    estilos = criar_estilos_resumo()

    ws.append(
        [
            "Nome",
            "Final da matrícula",
            "Setor",
            "Banco Total",
            "Banco Saldo",
            "Faltas",
        ]
    )

    for registro in dados:
        ws.append(
            [
                registro.nome,
                registro.final_matricula,
                registro.departamento,
                formatar_horas(registro.banco_total_minutos),
                formatar_horas(registro.banco_saldo_minutos),
                registro.faltas,
            ]
        )

    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.font = estilos["cabecalho_font"]
        cell.fill = estilos["cabecalho_fill"]
        cell.alignment = estilos["centro"]
        cell.border = estilos["borda"]

    for row in range(2, ws.max_row + 1):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = estilos["borda"]
            if col in (1, 3):
                cell.alignment = estilos["esquerda"]
            elif col == 2:
                cell.alignment = estilos["centro"]
                cell.number_format = "@"
            else:
                cell.alignment = estilos["direita"]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{ws.max_row}"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 12
