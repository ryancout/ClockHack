"""Estilos compartilhados pelas abas auxiliares do relatório."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def criar_estilos_resumo():
    """Cria os estilos usados nas abas SALDO, RANKING e RESUMO."""
    return {
        "titulo": Font(size=14, bold=True),
        "cabecalho_font": Font(bold=True, color="FFFFFF"),
        "cabecalho_fill": PatternFill("solid", fgColor="1F4E78"),
        "linha_total_font": Font(bold=True),
        "linha_total_fill": PatternFill("solid", fgColor="D9EAF7"),
        "borda": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
        "centro": Alignment(horizontal="center", vertical="center"),
        "direita": Alignment(horizontal="right", vertical="center"),
        "esquerda": Alignment(horizontal="left", vertical="center"),
    }
