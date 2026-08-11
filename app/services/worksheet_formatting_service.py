from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter


def formatar_aba_principal(ws):
    """Aplica a formatação visual da aba principal."""
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alinhamento = Alignment(vertical="center")

    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.border = borda
            cell.alignment = alinhamento

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)

        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column].width = max_length + 2
