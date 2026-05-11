from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from app.services.time_service import formatar_horas, para_minutos


def obter_ultima_linha(ws, col_nome):
    for row in range(ws.max_row, 1, -1):
        if ws.cell(row=row, column=col_nome).value not in (None, ""):
            return row
    return 1


def estilizar_linha_total(ws, nova_linha, col_inicio, col_bs):
    bold_font = Font(bold=True)
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=nova_linha, column=col)
        cell.border = borda
        if col_inicio <= col <= col_bs:
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=nova_linha, column=col_inicio).alignment = Alignment(horizontal="right", vertical="center")


def destacar_linhas_por_banco_saldo(ws, col_bs, ultima_linha_dados):
    vermelho = PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")
    amarelo = PatternFill(fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00")

    for row in range(2, ultima_linha_dados + 1):
        valor = ws.cell(row=row, column=col_bs).value
        if valor in (None, ""):
            continue

        try:
            total_min = para_minutos(valor)
        except Exception:
            continue

        fill = None
        if total_min < -480:
            fill = vermelho
        elif total_min > 480:
            fill = amarelo

        if fill:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill


def adicionar_legenda_cores(ws):
    vermelho = PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")
    amarelo = PatternFill(fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00")
    titulo_font = Font(bold=True)
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    col_legenda = ws.max_column + 2

    ws.cell(row=1, column=col_legenda, value="LEGENDA DE CORES")
    ws.cell(row=1, column=col_legenda).font = titulo_font

    ws.cell(row=2, column=col_legenda, value="Devedores")
    ws.cell(row=2, column=col_legenda + 1, value="Saldo menor que -8h")
    ws.cell(row=2, column=col_legenda).fill = vermelho

    ws.cell(row=3, column=col_legenda, value="Extras")
    ws.cell(row=3, column=col_legenda + 1, value="Saldo maior que 8h")
    ws.cell(row=3, column=col_legenda).fill = amarelo

    for row in range(1, 4):
        for col in range(col_legenda, col_legenda + 2):
            cell = ws.cell(row=row, column=col)
            cell.border = borda
            cell.alignment = Alignment(horizontal="center" if row == 1 else "left", vertical="center")

    ws.column_dimensions[ws.cell(row=1, column=col_legenda).column_letter].width = 18
    ws.column_dimensions[ws.cell(row=1, column=col_legenda + 1).column_letter].width = 22


def escrever_resultado(ws, col_nome, col_bt, col_bs, soma_bt, soma_bs):
    ultima_linha = obter_ultima_linha(ws, col_nome)
    destacar_linhas_por_banco_saldo(ws, col_bs, ultima_linha)

    nova_linha = ultima_linha + 1
    ws.cell(row=nova_linha, column=col_bt, value=formatar_horas(soma_bt))
    ws.cell(row=nova_linha, column=col_bs, value=formatar_horas(soma_bs))

    col_inicio = max(1, col_bt - 3)
    ws.merge_cells(
        start_row=nova_linha,
        start_column=col_inicio,
        end_row=nova_linha,
        end_column=col_bt - 1,
    )
    ws.cell(row=nova_linha, column=col_inicio, value="TOTAL")
    estilizar_linha_total(ws, nova_linha, col_inicio, col_bs)
    adicionar_legenda_cores(ws)
