import os
<<<<<<< HEAD
from collections import OrderedDict
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
=======
from openpyxl.chart import BarChart, Reference
>>>>>>> b23b4f0fc185652037e9b32f404393c6f1acc595

from app.services.reader_service import carregar_workbook
from app.services.validator_service import mapear_colunas, validar_colunas, validar_resultado
from app.services.filter_service import aplicar_filtro_departamento, listar_departamentos
from app.services.calculator_service import calcular_totais
from app.services.writer_service import escrever_resultado
from app.services.time_service import formatar_horas, para_minutos


def obter_departamentos(caminho_arquivo):
    try:
        wb = carregar_workbook(caminho_arquivo)
    except Exception as e:
        raise Exception(f"Erro ao abrir a planilha: {e}")

    try:
        ws = wb.active
    except Exception as e:
        raise Exception(f"Erro ao acessar a aba ativa: {e}")

    try:
        colunas = mapear_colunas(ws)
    except Exception as e:
        raise Exception(f"Erro ao mapear colunas: {e}")

    try:
        validar_colunas(colunas)
    except Exception as e:
        raise Exception(f"Erro na validação das colunas: {e}")

    try:
        return listar_departamentos(ws, colunas["nome do departamento"])
    except Exception as e:
        raise Exception(f"Erro ao listar departamentos: {e}")


<<<<<<< HEAD
def _estilos_resumo():
    borda = Border(
        left=Side(style="thin", color="D5DFE8"),
        right=Side(style="thin", color="D5DFE8"),
        top=Side(style="thin", color="D5DFE8"),
        bottom=Side(style="thin", color="D5DFE8"),
    )
    return {
        "titulo": Font(bold=True, size=15),
        "cabecalho_font": Font(bold=True, color="FFFFFF"),
        "cabecalho_fill": PatternFill(fill_type="solid", start_color="17324D", end_color="17324D"),
        "linha_total_font": Font(bold=True),
        "linha_total_fill": PatternFill(fill_type="solid", start_color="EAF2FB", end_color="EAF2FB"),
        "borda": borda,
        "centro": Alignment(horizontal="center", vertical="center"),
        "direita": Alignment(horizontal="right", vertical="center"),
        "esquerda": Alignment(horizontal="left", vertical="center"),
    }


def criar_aba_ranking(wb, dados):
=======
def criar_aba_ranking(wb, dados):
    """
    Cria uma aba com:
    - TOP devedores (mais saldo negativo)
    - TOP hora extra (mais saldo positivo)
    """
    # se já existir, remove e recria
>>>>>>> b23b4f0fc185652037e9b32f404393c6f1acc595
    if "RANKING" in wb.sheetnames:
        del wb["RANKING"]

    ws = wb.create_sheet("RANKING")
<<<<<<< HEAD
    estilos = _estilos_resumo()

    negativos = sorted([d for d in dados if d["saldo"] < 0], key=lambda x: x["saldo"])
    positivos = sorted([d for d in dados if d["saldo"] > 0], key=lambda x: x["saldo"], reverse=True)

    ws["A1"] = "TOP DEVEDORES"
    ws["A1"].font = estilos["titulo"]
    ws.append(["Funcionário", "Departamento", "Banco Saldo"])
    for d in negativos[:10]:
        ws.append([d["nome"], d["departamento"], d["saldo_fmt"]])

    inicio_segunda_secao = ws.max_row + 3
    ws.cell(row=inicio_segunda_secao, column=1, value="TOP HORAS EXTRAS")
    ws.cell(row=inicio_segunda_secao, column=1).font = estilos["titulo"]
    ws.append(["Funcionário", "Departamento", "Banco Saldo"])
    for d in positivos[:10]:
        ws.append([d["nome"], d["departamento"], d["saldo_fmt"]])

    for row in (2, inicio_segunda_secao + 1):
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.font = estilos["cabecalho_font"]
            cell.fill = estilos["cabecalho_fill"]
            cell.alignment = estilos["centro"]
            cell.border = estilos["borda"]

    for row in range(3, ws.max_row + 1):
        if row == inicio_segunda_secao:
            continue
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.border = estilos["borda"]
            cell.alignment = estilos["esquerda"] if col < 3 else estilos["direita"]

    ws.freeze_panes = None
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16


def criar_aba_resumo(wb, dados):
=======

    # separar dados
    negativos = sorted(
        [d for d in dados if d["saldo"] < 0],
        key=lambda x: x["saldo"]
    )

    positivos = sorted(
        [d for d in dados if d["saldo"] > 0],
        key=lambda x: x["saldo"],
        reverse=True
    )

    linha = 1

    # título 1
    ws.cell(row=linha, column=1, value="TOP DEVEDORES")
    linha += 1

    ws.append(["Funcionário", "Departamento", "Banco Saldo"])
    linha += 1

    for d in negativos[:10]:
        ws.append([d["nome"], d["departamento"], d["saldo_fmt"]])
        linha += 1

    linha += 2

    # título 2
    ws.cell(row=linha, column=1, value="TOP HORAS EXTRAS")
    linha += 1

    ws.append(["Funcionário", "Departamento", "Banco Saldo"])
    linha += 1

    for d in positivos[:10]:
        ws.append([d["nome"], d["departamento"], d["saldo_fmt"]])

    # ajustar largura básica
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15


def criar_aba_resumo(wb, dados):
    """
    Cria uma aba RESUMO com:
    - total de horas por departamento
    - gráfico automático
    """
    # se já existir, remove e recria
>>>>>>> b23b4f0fc185652037e9b32f404393c6f1acc595
    if "RESUMO" in wb.sheetnames:
        del wb["RESUMO"]

    ws = wb.create_sheet("RESUMO")
<<<<<<< HEAD
    estilos = _estilos_resumo()

    resumo = OrderedDict()
    for d in sorted(dados, key=lambda item: str(item["departamento"] or "SEM DEPARTAMENTO").lower()):
=======

    resumo = {}

    for d in dados:
>>>>>>> b23b4f0fc185652037e9b32f404393c6f1acc595
        departamento = d["departamento"] if d["departamento"] not in (None, "") else "SEM DEPARTAMENTO"
        resumo.setdefault(departamento, 0)
        resumo[departamento] += d["saldo"]

<<<<<<< HEAD
    ws["A1"] = "Resumo por departamento"
    ws["A1"].font = estilos["titulo"]
    ws.merge_cells("A1:C1")
    ws["A1"].alignment = estilos["esquerda"]

    ws.append(["Departamento", "Horas", "Horas_num"])

    for departamento, total_min in resumo.items():
        ws.append([departamento, formatar_horas(total_min), total_min / 60])

    linha_total = ws.max_row + 1
    total_geral_min = sum(resumo.values())
    ws.cell(row=linha_total, column=1, value="TOTAL")
    ws.cell(row=linha_total, column=2, value=formatar_horas(total_geral_min))
    ws.cell(row=linha_total, column=3, value=total_geral_min / 60)

    for col in range(1, 4):
        cell = ws.cell(row=2, column=col)
        cell.font = estilos["cabecalho_font"]
        cell.fill = estilos["cabecalho_fill"]
        cell.alignment = estilos["centro"]
        cell.border = estilos["borda"]

    for row in range(3, linha_total + 1):
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.border = estilos["borda"]
            if col == 1:
                cell.alignment = estilos["esquerda"]
            else:
                cell.alignment = estilos["direita"]

    for col in range(1, 4):
        cell = ws.cell(row=linha_total, column=col)
        cell.font = estilos["linha_total_font"]
        cell.fill = estilos["linha_total_fill"]
        cell.border = estilos["borda"]

    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col in range(1, ws.max_column + 1):
        ws.cell(row=linha_total, column=col).border = borda

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:B{linha_total}"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].hidden = True

=======
    ws.append(["Departamento", "Horas"])

    for departamento, total_min in resumo.items():
        total_horas = total_min / 60
        ws.append([departamento, total_horas])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15

    # gráfico
    if len(resumo) > 0:
        chart = BarChart()
        chart.title = "Horas por Departamento"
        chart.y_axis.title = "Horas"
        chart.x_axis.title = "Departamento"
        chart.height = 8
        chart.width = 16

        data = Reference(ws, min_col=2, min_row=1, max_row=len(resumo) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(resumo) + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "D2")
>>>>>>> b23b4f0fc185652037e9b32f404393c6f1acc595


def processar_arquivo(caminho_arquivo, caminho_saida, departamento="Todos"):
    wb = carregar_workbook(caminho_arquivo)
    ws = wb.active

    colunas = mapear_colunas(ws)
    validar_colunas(colunas)

    col_nome = colunas["nome do funcionário"]
    col_depart = colunas["nome do departamento"]
    col_bt = colunas["banco total"]
    col_bs = colunas["banco saldo"]

<<<<<<< HEAD
    aplicar_filtro_departamento(ws, col_depart, departamento)

=======
    # aplica filtro por departamento antes de processar
    aplicar_filtro_departamento(ws, col_depart, departamento)

    # coleta dados para ranking e resumo
>>>>>>> b23b4f0fc185652037e9b32f404393c6f1acc595
    dados = []
    for row in range(2, ws.max_row + 1):
        nome = ws.cell(row=row, column=col_nome).value
        dep = ws.cell(row=row, column=col_depart).value
        saldo_valor = ws.cell(row=row, column=col_bs).value

        if nome in (None, ""):
            continue

        saldo_min = para_minutos(saldo_valor)
<<<<<<< HEAD
=======

>>>>>>> b23b4f0fc185652037e9b32f404393c6f1acc595
        dados.append({
            "nome": nome,
            "departamento": dep,
            "saldo": saldo_min,
<<<<<<< HEAD
            "saldo_fmt": formatar_horas(saldo_min),
        })

    resultado_calc = calcular_totais(ws, col_nome, col_bt, col_bs)
    validar_resultado(resultado_calc["quantidade_funcionarios"])

=======
            "saldo_fmt": formatar_horas(saldo_min)
        })

    # cálculos
    resultado_calc = calcular_totais(ws, col_nome, col_bt, col_bs)

    validar_resultado(resultado_calc["quantidade_funcionarios"])

    # escreve linha TOTAL na aba principal
>>>>>>> b23b4f0fc185652037e9b32f404393c6f1acc595
    escrever_resultado(
        ws,
        col_nome,
        col_bt,
        col_bs,
        resultado_calc["soma_bt"],
<<<<<<< HEAD
        resultado_calc["soma_bs"],
    )

    criar_aba_ranking(wb, dados)
    criar_aba_resumo(wb, dados)

=======
        resultado_calc["soma_bs"]
    )

    # PASSO 4 entra aqui:
    # criar abas extras antes de salvar
    criar_aba_ranking(wb, dados)
    criar_aba_resumo(wb, dados)

    # salva o arquivo
>>>>>>> b23b4f0fc185652037e9b32f404393c6f1acc595
    wb.save(caminho_saida)

    return {
        "caminho_saida": caminho_saida,
        "banco_total": formatar_horas(resultado_calc["soma_bt"]),
        "banco_saldo": formatar_horas(resultado_calc["soma_bs"]),
        "quantidade_funcionarios": resultado_calc["quantidade_funcionarios"],
        "tipo_entrada": os.path.splitext(caminho_arquivo)[1].lower().replace(".", "").upper(),
<<<<<<< HEAD
        "departamento": departamento,
    }
=======
        "departamento": departamento
    }
>>>>>>> b23b4f0fc185652037e9b32f404393c6f1acc595
