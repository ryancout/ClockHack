"""Prepara relatórios do FAS Jornada para consumo analítico no Power BI."""

from __future__ import annotations

import re
import unicodedata
import uuid
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.domain import para_minutos


class AnalyticsError(Exception):
    """Relatório incompatível com a exportação analítica."""


@dataclass(frozen=True, slots=True)
class PowerBiSnapshot:
    report_id: str
    rows: tuple[dict[str, Any], ...]
    source_file: str

    @property
    def row_count(self) -> int:
        return len(self.rows)


_CAMPOS_HORAS = {
    "TotalNormaisMinutos": ("total normais",),
    "TotalTrabalhadoMinutos": ("total trabalhado",),
    "HorasPrevistasMinutos": ("horas previstas",),
    "FaltaAtrasoMinutos": ("falta e atraso",),
    "AbonoMinutos": ("abono",),
    "ExtrasTotalMinutos": ("extras total", "extra total"),
    "BancoTotalMinutos": ("banco total",),
    "BancoSaldoMinutos": ("banco saldo",),
}


def _normalizar(valor: object) -> str:
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    texto = "".join(char for char in texto if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", texto).strip().casefold()


def _texto(valor: object) -> str:
    return str(valor or "").strip()


def _inteiro(valor: object) -> int:
    if valor in (None, ""):
        return 0
    try:
        return int(float(str(valor).replace(",", ".")))
    except (TypeError, ValueError):
        return 0


def _periodo(resultado: dict[str, Any]) -> tuple[str | None, str | None]:
    inicial = resultado.get("periodo_inicial")
    final = resultado.get("periodo_final")
    if inicial or final:
        return _iso_data(inicial), _iso_data(final)

    periodo = _texto(resultado.get("periodo_rhid"))
    encontrados = re.findall(r"\d{4}-\d{2}-\d{2}", periodo)
    if len(encontrados) >= 2:
        return encontrados[0], encontrados[1]
    return None, None


def _iso_data(valor: object) -> str | None:
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    texto = _texto(valor)
    try:
        return date.fromisoformat(texto).isoformat() if texto else None
    except ValueError:
        return None


def _classificar_saldo(minutos: int) -> str:
    if minutos < -480:
        return "Saldo crítico"
    if minutos < 0:
        return "Saldo negativo"
    if minutos == 0:
        return "Equilibrado"
    if minutos <= 480:
        return "Saldo positivo"
    return "Excesso de horas"


def _indice(cabecalhos: Sequence[object], *nomes: str) -> int | None:
    mapa = {_normalizar(valor): indice for indice, valor in enumerate(cabecalhos)}
    for nome in nomes:
        if nome in mapa:
            return mapa[nome]
    return None


def preparar_snapshot_powerbi(
    caminho_xlsx: str | Path,
    resultado: dict[str, Any],
    *,
    report_id: str | None = None,
    generated_at: datetime | None = None,
) -> PowerBiSnapshot:
    """Extrai uma linha analítica por funcionário, sem incluir CPF."""

    caminho = Path(caminho_xlsx)
    if not caminho.is_file():
        raise AnalyticsError("O arquivo do relatório não foi encontrado.")

    workbook = load_workbook(caminho, read_only=True, data_only=True)
    try:
        worksheet = workbook["Sheet"] if "Sheet" in workbook.sheetnames else workbook.active
        if worksheet is None:
            raise AnalyticsError("O relatório não possui uma planilha principal.")
        linhas = worksheet.iter_rows(values_only=True)
        cabecalhos = list(next(linhas, ()))
        indice_nome = _indice(cabecalhos, "nome do funcionario", "funcionario", "nome")
        indice_matricula = _indice(
            cabecalhos,
            "numero de matricula",
            "matricula",
            "final da matricula",
        )
        indice_departamento = _indice(
            cabecalhos, "nome do departamento", "departamento", "setor"
        )
        if indice_nome is None or indice_matricula is None or indice_departamento is None:
            raise AnalyticsError(
                "O relatório não possui nome, matrícula e departamento."
            )

        indices_horas = {
            destino: _indice(cabecalhos, *aliases)
            for destino, aliases in _CAMPOS_HORAS.items()
        }
        indice_faltas = _indice(cabecalhos, "dia falta", "faltas")
        identificador = report_id or str(uuid.uuid4())
        gerado_em = (generated_at or datetime.now().astimezone()).isoformat()
        periodo_inicial, periodo_final = _periodo(resultado)
        competencia_base = periodo_final or gerado_em[:10]
        competencia = competencia_base[:7]
        ano = int(competencia_base[:4])
        mes = int(competencia_base[5:7])
        origem = _texto(resultado.get("tipo_entrada") or "CSV").upper()
        empresa = _texto(resultado.get("empresa") or resultado.get("company_label"))
        setores = _texto(
            resultado.get("setores")
            or resultado.get("department_label")
            or resultado.get("departamento")
        )

        saida: list[dict[str, Any]] = []
        for valores in linhas:
            nome = _texto(valores[indice_nome] if indice_nome < len(valores) else None)
            matricula = _texto(
                valores[indice_matricula] if indice_matricula < len(valores) else None
            )
            if not nome or not matricula or _normalizar(nome) == "total":
                continue
            departamento = _texto(
                valores[indice_departamento]
                if indice_departamento < len(valores)
                else None
            )
            minutos: dict[str, int] = {}
            for destino, indice_coluna in indices_horas.items():
                valor = (
                    valores[indice_coluna]
                    if indice_coluna is not None and indice_coluna < len(valores)
                    else None
                )
                minutos[destino] = para_minutos(valor)

            trabalhado = minutos["TotalTrabalhadoMinutos"]
            previsto = minutos["HorasPrevistasMinutos"]
            saldo = minutos["BancoSaldoMinutos"]
            diferenca = trabalhado - previsto
            classificacao = _classificar_saldo(saldo)
            linha = {
                "IDRelatorio": identificador,
                "GeradoEm": gerado_em,
                "PeriodoInicial": periodo_inicial,
                "PeriodoFinal": periodo_final,
                "Competencia": competencia,
                "Ano": ano,
                "Mes": mes,
                "Origem": origem,
                "Empresa": empresa,
                "SetoresSelecionados": setores,
                "Arquivo": caminho.name,
                "Matricula": matricula,
                "Funcionario": nome,
                "Departamento": departamento,
                **minutos,
                "FaltasDias": _inteiro(
                    valores[indice_faltas]
                    if indice_faltas is not None and indice_faltas < len(valores)
                    else None
                ),
                "TotalTrabalhadoHoras": round(trabalhado / 60, 4),
                "TotalNormaisHoras": round(minutos["TotalNormaisMinutos"] / 60, 4),
                "HorasPrevistasHoras": round(previsto / 60, 4),
                "FaltaAtrasoHoras": round(minutos["FaltaAtrasoMinutos"] / 60, 4),
                "AbonoHoras": round(minutos["AbonoMinutos"] / 60, 4),
                "ExtrasTotalHoras": round(minutos["ExtrasTotalMinutos"] / 60, 4),
                "BancoTotalHoras": round(minutos["BancoTotalMinutos"] / 60, 4),
                "BancoSaldoHoras": round(saldo / 60, 4),
                "DiferencaJornadaMinutos": diferenca,
                "DiferencaJornadaHoras": round(diferenca / 60, 4),
                "PercentualTrabalhado": (
                    round((trabalhado / previsto) * 100, 2) if previsto else 0.0
                ),
                "ClassificacaoSaldo": classificacao,
                "ClassificacaoSaldoOrdem": {
                    "Saldo crítico": 1,
                    "Saldo negativo": 2,
                    "Equilibrado": 3,
                    "Saldo positivo": 4,
                    "Excesso de horas": 5,
                }[classificacao],
                "IndicadorSaldoCritico": int(abs(saldo) > 480),
            }
            if periodo_inicial is None:
                linha.pop("PeriodoInicial")
            if periodo_final is None:
                linha.pop("PeriodoFinal")
            saida.append(linha)
    finally:
        workbook.close()

    if not saida:
        raise AnalyticsError("O relatório não possui funcionários para enviar.")
    return PowerBiSnapshot(identificador, tuple(saida), str(caminho))


__all__ = ["AnalyticsError", "PowerBiSnapshot", "preparar_snapshot_powerbi"]
