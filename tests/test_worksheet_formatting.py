from openpyxl import Workbook

from app.services.worksheet_formatting_service import formatar_aba_principal


def test_formatar_aba_principal_aplica_estilo_e_largura():
    wb = Workbook()
    ws = wb.active
    ws.append(["Nome", "Saldo"])
    ws.append(["Funcionário Teste", "1:00"])

    formatar_aba_principal(ws)

    assert ws["A1"].border.left.style == "thin"
    assert ws["A1"].border.right.style == "thin"
    assert ws["A1"].alignment.vertical == "center"
    assert ws.column_dimensions["A"].width == len("Funcionário Teste") + 2
    assert ws.column_dimensions["B"].width == len("Saldo") + 2
