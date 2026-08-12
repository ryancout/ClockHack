from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.integrations.rhid_client import RhidApiError
from app.integrations.rhid_report_service import RhidReportPlan, processar_relatorio_rhid


FIXTURE = Path(__file__).parent / "fixtures" / "relatorio_contrato_anonimo.csv"


class FakeRhidClient:
    def __init__(self, conteudo):
        self.conteudo = conteudo
        self.chamada = None

    def gerar_relatorio_csv(
        self,
        company_id,
        department_id,
        data_inicial,
        data_final,
        ao_progresso,
    ):
        self.chamada = (company_id, department_id, data_inicial, data_final)
        ao_progresso(50)
        ao_progresso(100)
        return self.conteudo


def test_relatorio_rhid_reutiliza_pipeline_e_remove_csv_temporario(tmp_path):
    cliente = FakeRhidClient(FIXTURE.read_bytes())
    saida = tmp_path / "relatorio_rhid.xlsx"
    plano = RhidReportPlan(
        company_id=7,
        department_id=12,
        company_label="Projeto A",
        department_label="Operações",
        data_inicial=date(2026, 8, 1),
        data_final=date(2026, 8, 11),
        caminho_saida=str(saida),
    )
    progresso = []

    resultado = processar_relatorio_rhid(cliente, plano, progresso.append)

    assert saida.exists()
    assert cliente.chamada == (7, 12, date(2026, 8, 1), date(2026, 8, 11))
    assert resultado["tipo_entrada"] == "RHID"
    assert resultado["departamento"] == "Operações"
    assert resultado["empresa_rhid"] == "Projeto A"
    assert progresso == [40, 80, 85, 100]
    assert not list(tmp_path.glob(".clockhack_rhid_*.csv"))

    workbook = load_workbook(saida, read_only=True)
    assert workbook.sheetnames == ["Sheet", "SALDO", "RANKING", "RESUMO"]


def test_relatorio_rhid_respeita_abas_opcionais(tmp_path):
    cliente = FakeRhidClient(FIXTURE.read_bytes())
    saida = tmp_path / "relatorio_rhid_sem_abas.xlsx"
    plano = RhidReportPlan(
        company_id=7,
        department_id=(12, 13),
        company_label="Projeto A",
        department_label="2 setores selecionados",
        data_inicial=date(2026, 8, 1),
        data_final=date(2026, 8, 11),
        caminho_saida=str(saida),
        gerar_saldo=False,
        gerar_resumo=False,
        gerar_ranking=False,
    )

    resultado = processar_relatorio_rhid(cliente, plano)

    assert cliente.chamada[1] == (12, 13)
    assert resultado["gerou_saldo"] is False
    assert resultado["gerou_resumo"] is False
    assert resultado["gerou_ranking"] is False
    workbook = load_workbook(saida, read_only=True)
    assert workbook.sheetnames == ["Sheet"]


def test_relatorio_rhid_recusa_matricula_duplicada_sem_somar_saldos(tmp_path):
    cabecalho = (
        "Nome do funcionário;Número de matrícula;Nome do departamento;"
        "CPF do funcionário;Total Normais;Total Trabalhado;Horas Previstas;"
        "Dia Falta;Falta e Atraso;Abono;Extras Total;Banco Total;Banco Saldo\r\n"
    )
    conteudo = (
        cabecalho
        + "Pessoa;MAT-0001;Operações;11111111111;08:00;08:00;08:00;0;0:00;0:00;0:00;1:00;2:00\r\n"
        + "Pessoa;MAT-0001;Operações;11111111111;08:00;08:00;08:00;0;0:00;0:00;0:00;-0:30;1:30\r\n"
    ).encode("utf-8-sig")
    cliente = FakeRhidClient(conteudo)
    saida = tmp_path / "nao_deve_existir.xlsx"
    plano = RhidReportPlan(
        company_id=7,
        department_id=12,
        company_label="Projeto A",
        department_label="Operações",
        data_inicial=date(2026, 8, 1),
        data_final=date(2026, 8, 11),
        caminho_saida=str(saida),
    )

    with pytest.raises(RhidApiError, match="mais de uma linha"):
        processar_relatorio_rhid(cliente, plano)

    assert not saida.exists()
    assert not list(tmp_path.glob(".clockhack_rhid_*.csv"))


def test_relatorio_rhid_recusa_formato_diario_de_inconsistencias(tmp_path):
    conteudo = (
        "Nome do funcionário;Número de matrícula;Nome do departamento;"
        "CPF do funcionário;Dia;Total Normais;Total Trabalhado;Banco Total;Banco Saldo\r\n"
        "Pessoa;MAT-0001;Operações;11111111111;2026-08-01;08:00;08:00;1:00;2:00\r\n"
    ).encode("utf-8-sig")
    cliente = FakeRhidClient(conteudo)
    saida = tmp_path / "nao_deve_existir.xlsx"
    plano = RhidReportPlan(
        company_id=7,
        department_id=None,
        company_label="Projeto A",
        department_label="Todos os setores",
        data_inicial=date(2026, 8, 1),
        data_final=date(2026, 8, 11),
        caminho_saida=str(saida),
    )

    with pytest.raises(RhidApiError, match="relatório consolidado"):
        processar_relatorio_rhid(cliente, plano)

    assert not saida.exists()
