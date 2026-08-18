import csv
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.core.exceptions import ArquivoInvalidoError, ColunaObrigatoriaError
from app.services.validator_service import validar_arquivo_entrada
from app.services.workbook_pipeline_service import obter_departamentos, processar_arquivo


FIXTURE_CSV = Path(__file__).parent / "fixtures" / "relatorio_contrato_anonimo.csv"

CPFS_FICTICIOS = {
    "000.000.000-00",
    "111.111.111-11",
    "222.222.222-22",
    "333.333.333-33",
}

CABECALHO_ESPERADO = (
    "Nome do funcionário",
    "Número de matrícula",
    "Nome do departamento",
    "CPF do funcionário",
    "Total Normais",
    "Total Trabalhado",
    "Horas Previstas",
    "Dia Falta",
    "Falta e Atraso",
    "Abono",
    "Extras Total",
    "Banco Total",
    "Banco Saldo",
)


def test_fixture_reproduz_contrato_csv_sem_dados_reais():
    conteudo = FIXTURE_CSV.read_bytes()
    assert conteudo.startswith(b"\xef\xbb\xbf")

    with FIXTURE_CSV.open(encoding="utf-8-sig", newline="") as arquivo:
        linhas = list(csv.reader(arquivo, delimiter=";"))

    assert tuple(linhas[0]) == CABECALHO_ESPERADO
    assert len(linhas) == 5
    assert all(len(linha) == 13 for linha in linhas)

    funcionarios = linhas[1:]
    homonimos = [(linha[0], linha[1]) for linha in funcionarios if linha[0] == "Ana Teste"]
    assert homonimos == [("Ana Teste", "MAT-0001"), ("Ana Teste", "MAT-0002")]

    assert funcionarios[0][11] == "26:30"
    assert funcionarios[1][12] == "-10:30"
    assert funcionarios[2][11] == ""
    assert funcionarios[3][12] == ""
    assert {linha[3] for linha in funcionarios} == {
        "000.000.000-00",
        "111.111.111-11",
        "222.222.222-22",
        "333.333.333-33",
    }


def test_validacao_aceita_csv_e_rejeita_xlsx_existente(tmp_path):
    validar_arquivo_entrada(str(FIXTURE_CSV))

    caminho_xlsx = tmp_path / "entrada.xlsx"
    Workbook().save(caminho_xlsx)

    with pytest.raises(ArquivoInvalidoError, match="CSV"):
        validar_arquivo_entrada(str(caminho_xlsx))

    with pytest.raises(ArquivoInvalidoError, match="CSV"):
        processar_arquivo(str(caminho_xlsx), str(tmp_path / "saida.xlsx"))


def test_processamento_exige_cabecalho_na_primeira_linha(tmp_path):
    caminho_csv = tmp_path / "cabecalho_deslocado.csv"
    conteudo = FIXTURE_CSV.read_text(encoding="utf-8-sig")
    caminho_csv.write_text(
        "Relatório de teste\n" + conteudo,
        encoding="utf-8-sig",
    )

    with pytest.raises(ArquivoInvalidoError, match="primeira linha"):
        processar_arquivo(str(caminho_csv), str(tmp_path / "saida.xlsx"))


def test_processamento_exige_numero_de_matricula(tmp_path):
    caminho_csv = tmp_path / "sem_matricula.csv"
    conteudo = FIXTURE_CSV.read_text(encoding="utf-8-sig").replace(
        "Número de matrícula",
        "Identificador removido",
        1,
    )
    caminho_csv.write_text(conteudo, encoding="utf-8-sig")

    with pytest.raises(ColunaObrigatoriaError, match="Número de matrícula"):
        processar_arquivo(str(caminho_csv), str(tmp_path / "saida.xlsx"))


def test_processamento_preserva_homonimos_totais_e_abas(tmp_path):
    caminho_saida = tmp_path / "resultado.xlsx"

    resultado = processar_arquivo(str(FIXTURE_CSV), str(caminho_saida))

    assert resultado["quantidade_funcionarios"] == 4
    assert resultado["banco_total"] == "24:15"
    assert resultado["banco_saldo"] == "0:00"
    assert resultado["tipo_entrada"] == "CSV"

    wb = load_workbook(caminho_saida)
    assert wb.sheetnames[1:] == ["SALDO", "RANKING", "RESUMO"]

    ws_principal = wb.worksheets[0]
    cabecalhos = {
        ws_principal.cell(row=1, column=coluna).value: coluna
        for coluna in range(1, ws_principal.max_column + 1)
    }
    col_nome = cabecalhos["Nome do funcionário"]
    assert "CPF do funcionário" not in cabecalhos
    col_matricula = cabecalhos["Número de matrícula"]
    col_banco_total = cabecalhos["Banco Total"]
    col_banco_saldo = cabecalhos["Banco Saldo"]

    homonimos = [
        (
            ws_principal.cell(row=linha, column=col_nome).value,
            ws_principal.cell(row=linha, column=col_matricula).value,
        )
        for linha in range(2, 6)
        if ws_principal.cell(row=linha, column=col_nome).value == "Ana Teste"
    ]
    assert homonimos == [("Ana Teste", "MAT-0001"), ("Ana Teste", "MAT-0002")]

    matriculas_principal = [
        ws_principal.cell(row=linha, column=col_matricula).value
        for linha in range(2, 6)
    ]
    assert matriculas_principal == ["MAT-0001", "MAT-0002", "MAT-0003", "MAT-0004"]
    assert all(isinstance(matricula, str) for matricula in matriculas_principal)
    assert ws_principal.cell(row=6, column=col_banco_total).value == "24:15"
    assert ws_principal.cell(row=6, column=col_banco_saldo).value == "0:00"

    ws_saldo = wb["SALDO"]
    assert [cell.value for cell in ws_saldo[1]] == [
        "Nome",
        "Número de matrícula",
        "Setor",
        "Banco Total",
        "Banco Saldo",
        "Faltas",
    ]

    matriculas_saldo = [ws_saldo.cell(row=linha, column=2).value for linha in range(2, 6)]
    assert matriculas_saldo == ["MAT-0001", "MAT-0002", "MAT-0003", "MAT-0004"]

    saldos_homonimos = [
        (
            ws_saldo.cell(row=linha, column=2).value,
            ws_saldo.cell(row=linha, column=4).value,
            ws_saldo.cell(row=linha, column=5).value,
        )
        for linha in range(2, ws_saldo.max_row + 1)
        if ws_saldo.cell(row=linha, column=1).value == "Ana Teste"
    ]
    assert saldos_homonimos == [
        ("MAT-0001", "26:30", "9:15"),
        ("MAT-0002", "-2:45", "-10:30"),
    ]

    ws_ranking = wb["RANKING"]
    assert sum(cell.value == "Ana Teste" for cell in ws_ranking["A"]) == 2
    cabecalhos_ranking = [
        [cell.value for cell in ws_ranking[linha]]
        for linha in range(1, ws_ranking.max_row + 1)
        if ws_ranking.cell(row=linha, column=1).value == "Funcionário"
    ]
    assert cabecalhos_ranking == [
        ["Funcionário", "Número de matrícula", "Departamento", "Banco Saldo"],
        ["Funcionário", "Número de matrícula", "Departamento", "Banco Saldo"],
    ]

    ranking_homonimos = sorted(
        (
            ws_ranking.cell(row=linha, column=2).value,
            ws_ranking.cell(row=linha, column=4).value,
        )
        for linha in range(1, ws_ranking.max_row + 1)
        if ws_ranking.cell(row=linha, column=1).value == "Ana Teste"
    )
    assert ranking_homonimos == [("MAT-0001", "9:15"), ("MAT-0002", "-10:30")]

    for ws_auxiliar in wb.worksheets:
        valores = {
            str(cell.value)
            for linha in ws_auxiliar.iter_rows()
            for cell in linha
            if cell.value is not None
        }
        assert all(cpf not in valor for valor in valores for cpf in CPFS_FICTICIOS)
        assert all("cpf" not in valor.casefold() for valor in valores)

    ws_resumo = wb["RESUMO"]
    resumo = {
        ws_resumo.cell(row=linha, column=1).value: ws_resumo.cell(row=linha, column=2).value
        for linha in range(3, ws_resumo.max_row + 1)
        if ws_resumo.cell(row=linha, column=1).value in {"Financeiro", "Operações", "TOTAL"}
    }
    assert resumo == {"Operações": "-1:15", "Financeiro": "1:15", "TOTAL": "0:00"}


def test_filtro_por_departamento_mantem_os_dois_homonimos(tmp_path):
    assert obter_departamentos(str(FIXTURE_CSV)) == ["Todos", "Financeiro", "Operações"]

    caminho_saida = tmp_path / "resultado_operacoes.xlsx"
    resultado = processar_arquivo(
        str(FIXTURE_CSV),
        str(caminho_saida),
        departamento="Operações",
    )

    assert resultado["quantidade_funcionarios"] == 2
    assert resultado["banco_total"] == "23:45"
    assert resultado["banco_saldo"] == "-1:15"

    wb = load_workbook(caminho_saida)
    ws_principal = wb.worksheets[0]
    assert [ws_principal.cell(row=linha, column=1).value for linha in range(2, 4)] == [
        "Ana Teste",
        "Ana Teste",
    ]
    assert ws_principal.cell(row=1, column=2).value == "Número de matrícula"
    assert [ws_principal.cell(row=linha, column=2).value for linha in range(2, 4)] == [
        "MAT-0001",
        "MAT-0002",
    ]
    assert all(
        isinstance(ws_principal.cell(row=linha, column=2).value, str)
        for linha in range(2, 4)
    )
    assert ws_principal.cell(row=4, column=11).value == "23:45"
    assert ws_principal.cell(row=4, column=12).value == "-1:15"

    ws_saldo = wb["SALDO"]
    assert [ws_saldo.cell(row=linha, column=2).value for linha in range(2, 4)] == [
        "MAT-0001",
        "MAT-0002",
    ]


def test_processamento_respeita_abas_opcionais(tmp_path):
    caminho_saida = tmp_path / "resultado_sem_abas_auxiliares.xlsx"

    resultado = processar_arquivo(
        str(FIXTURE_CSV),
        str(caminho_saida),
        gerar_saldo=False,
        gerar_ranking=False,
        gerar_resumo=False,
    )

    assert resultado["banco_total"] == "24:15"
    assert resultado["banco_saldo"] == "0:00"
    assert resultado["gerou_saldo"] is False
    assert resultado["gerou_ranking"] is False
    assert resultado["gerou_resumo"] is False

    wb = load_workbook(caminho_saida)
    assert wb.sheetnames == ["Sheet"]
    assert wb.active.cell(row=1, column=2).value == "Número de matrícula"
